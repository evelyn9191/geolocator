# Geolocator script fills GPS coordinates to Excel sheet based on address. User puts
# names of columns containing street, city name, postal code and column for gps coordinates
# in Excel sheet to the script. The script searches for gps coordinates using Nominatim
# geolocator and writes the coordinates to separate output Excel file.

import os
import re

import pandas as pd
from geopy.geocoders import Nominatim
from geopy.extra.rate_limiter import RateLimiter
from shutil import copyfile
from openpyxl import load_workbook
import xlrd


def orig_file_check():
    """Get input file name. Check if it exists and if the format is right.

    :return: Name of the input file.
    """
    input_file = input("Put the Excel file to the same directory, where you put "
                       "the file you are currently running. What is the name "
                       "of the Excel file? (Write it as filename.extension) ")
    while os.path.exists(input_file) is False:
        print("Such file does not exist. Did you put it into a right directory? "
              "Is the name of the file right?")
        input_file = input("Try again: ")
    while input_file.lower().endswith('.xlsx') is False:
        print("I cannot process such file. Give me a file with .xlsx extension "
              "(Excel file of version 2007 and later.)")
        input_file = input("Try again: ")
    return input_file


def user_data():
    """Get input data.

    Get indexes of columns with required data (street names, city names, postal code and
    column where gps coordinates should be placed.

    :return: [dict] with input data.
    """
    street = input("What is the title of the column with street names? ")
    city = input("What is the title of the column with city names? ")
    postal = input("What is the title of the column with postal codes? ")
    gps = input("What is the title of the column where the GPS coordinates should go? ")
    return {"street_column": street,
            "city_column": city,
            "postal_column": postal,
            "gps_column": gps
            }


def correct_data_check(input_file, user_data):
    """Check if input data are correct.

    :param str input_file: Name of the file with input data.
    :param dict user_data: Indices of columns containing required data.

    :return Dictionary with correct input data.
    """
    df = pd.read_excel(input_file)
    for key, value in user_data.items():
        while value not in df:
            print("The input %s is not correct." % value)
            new_variable = input("Write it again: ")
            value = new_variable
            user_data[key] = value
    return user_data


def get_gps(input_file, user_data):
    """Process :input_file according to :user_data.

    Open the input file, set delay and entry rate limit for Geolocator,
    get GPS coordinates for each line, and remove parentheses from coordinates.

    :param dict user_data: Indices of columns containing required data.
    :param str input_file: Name of the file to process.
    """
    df = pd.read_excel(input_file)
    df.fillna('0', inplace=True)
    df[user_data['postal_column']] = df[user_data['postal_column']].astype(int)

    geolocator = Nominatim(user_agent="Scratch Wars Shops GPS Finder")
    geocode = RateLimiter(geolocator.geocode, min_delay_seconds=1, max_retries=2, error_wait_seconds=3.0,
                          swallow_exceptions=True, return_value_on_exception=None)

    count_rows = df.shape[0]
    all_rows = range(count_rows)
    gps_column_index = df.columns.get_loc(user_data['gps_column']) + 1

    copyfile(input_file, "gps_coordinates.xlsx")
    output_file = "gps_coordinates.xlsx"
    wb = load_workbook(output_file)
    ws = wb.active

    for row_number in all_rows:
        line = df.loc[row_number, [user_data["street_column"], user_data["city_column"],
                                   user_data["postal_column"]]]
        line = list(line)    # Makes the output into a list readable for Geolocator
        line[1] = re.sub(r"\d", "", line[1])    # Removes number in city column
        print(line)
        location = geocode(line)    # Give address as needed to be processed by Geolocator.
        if location is None:
            continue
        clean_gps = "{}, {}".format(location.latitude, location.longitude)    # Remove parentheses
        ws.cell(row=row_number+2, column=gps_column_index).value = clean_gps
    wb.save(output_file)
    print("GPS coordinates successfully saved to", output_file)


if __name__ == "__main__":
    input_file = orig_file_check()
    user_data = user_data()
    clean_user_data = correct_data_check(input_file=input_file, user_data=user_data)
    get_gps(input_file=input_file, user_data=clean_user_data)
